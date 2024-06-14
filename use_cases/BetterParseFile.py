from io import TextIOWrapper
from use_cases.InputMatrix import InputMatrix
from use_cases.ParseFileInterface import ParseFileInterface
import openpyxl


class BetterParseFile(ParseFileInterface):

    def parse_file(self, file: TextIOWrapper) -> InputMatrix:
        # Open xlsx file as workbook, then get the correctly named sheet or the first sheet
        workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)

        # get sheets
        reviewersheet = workbook[workbook.sheetnames[0]]
        appsheet = workbook[workbook.sheetnames[1]]

        # create empty InputMatrix
        inputmatrix = InputMatrix()

        # process reviewers
        ## Get reviewers
        id_to_reviewer_index = {}
        reviewer_expertises = []     # list of dicts. expertises[i] contains the expertise dictionary for reviewer i. Keys of dicts in lowercase
        reviewer_program_areas = []  # list of list[str]. reviewer_program_areas[i] contains the program areas of reviewer i. Program area strings stored in lowercase
        row = 2
        while str(reviewersheet.cell(row, 2).value) != "" and reviewersheet.cell(row, 2).value != None:
            # get reviewer name
            inputmatrix.reviewers.append(str(reviewersheet.cell(row, 2).value))
            # get reviewer ID and match to index
            id = int(reviewersheet.cell(row, 1).value)
            id_to_reviewer_index[id] = len(inputmatrix.reviewers) -1

            # get program area
            program_area = str(reviewersheet.cell(row, 3).value).lower()
            program_area_split = program_area.split(";")
            for i, string in enumerate(program_area_split):
                program_area_split[i] = string.strip()
            reviewer_program_areas.append(program_area_split)

            row += 1

        # get first column with no expertise
        last_expertise = 4
        expertise_names = []     # list of expertise names, used later as keys for expertise dicts. Names in lowercase
        while str(reviewersheet.cell(1, last_expertise).value) != "" and reviewersheet.cell(1, last_expertise).value != None:
            expertise_names.append(str(reviewersheet.cell(1, last_expertise).value).lower())
            last_expertise += 1

        # get expertise values for each reviewer
        # also setup inputmatrix matrix
        for i, reviewer in enumerate(inputmatrix.reviewers):
            # get expertise values
            dict = {}
            for j, col in enumerate(range(4, last_expertise)):
                value = str(reviewersheet.cell(i+2, col).value).lower().strip()
                if value == "no expertise" or value == "":
                    dict[expertise_names[j]] = 0
                elif value == 'low':
                    dict[expertise_names[j]] = 1
                elif value == 'med' or value == 'medium':
                    dict[expertise_names[j]] = 2
                elif value == 'high':
                    dict[expertise_names[j]] = 3
            reviewer_expertises.append(dict)

            # setup inputmatrix matrix
            inputmatrix.matrix.append([])

        # process app sheet
        # Get app names and app_program_areas
        app_program_areas = []
        app_expertises = []
        app_conflicts = []      # stored as strings, need to cast to int before using as keys
        row = 2
        while str(appsheet.cell(row, 1).value) != "" and appsheet.cell(row, 1).value != None:
            # get name
            inputmatrix.applicants.append(str(appsheet.cell(row, 1).value))

            # get program areas
            raw = str(appsheet.cell(row, 2).value).lower()
            split_values = raw.split(';')
            app_program_areas.append(split_values)

            # get expertise values
            raw = str(appsheet.cell(row, 3).value).lower()
            split_values = raw.split(';')
            app_expertises.append(split_values)

            # get conflicts
            raw = str(appsheet.cell(row, 4).value)
            split_values = raw.split(';')
            app_conflicts.append(split_values)

            row += 1
        
        # fill out matrix
        for i, reviewer in enumerate(inputmatrix.reviewers):
            for j, applicant in enumerate(inputmatrix.applicants):
                compat = self.calculate_compat(i, j, reviewer_expertises, reviewer_program_areas, app_expertises, app_program_areas)
                inputmatrix.matrix[i].append(compat)

        # fill out conflicts
        for j, applicant in enumerate(inputmatrix.applicants):
            for conflict in app_conflicts[j]:
                try:
                    id = int(conflict)
                    reviewer_index = id_to_reviewer_index[id]
                    inputmatrix.matrix[reviewer_index][j] = -1
                except ValueError:
                    print("invalid conflict: " + conflict)

        return inputmatrix
    
    def calculate_compat(self, reviewer: int, applicant: int, reviewer_expertises: list, reviewer_program_areas: list, app_expertises: list, app_program_areas: list):
        score = 0

        # calculate program area, if one matches, set score to 1
        for program in reviewer_program_areas[reviewer]:
            if program in app_program_areas[applicant]:
                score = 1

        # calculate expertise
        for expertise in app_expertises[applicant]:
            if expertise in reviewer_expertises[reviewer]:
                score += reviewer_expertises[reviewer][expertise]
            else:
                print("could not find expertise: " + expertise)

        print("compat between reviewer " + str(reviewer) + " and applicant " + str(applicant) + " was " + str(score))
        return score