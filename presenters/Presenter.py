import openpyxl.styles
from presenters.PresenterInterface import PresenterInterface
from ui.UIInterface import UIInterface
from use_cases.InputMatrix import InputMatrix
from use_cases.OutputMatrix import OutputMatrix
import openpyxl
from tempfile import NamedTemporaryFile
from tkinter import filedialog as fd
import shutil

class Presenter(PresenterInterface):
    
    def present_output(self, output_matrix: OutputMatrix, input_matrix: InputMatrix):
        # turn output matrix into a xlsx file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "pairings"

        # do assignment sheet stuff
        self.add_headers_to_sheet(sheet, output_matrix)
        # fill out the matrix
        for applicant, assignments in enumerate(output_matrix.applicant_assignments):
            #print(assignments)
            for rank, reviewer in enumerate(assignments):
                sheet.cell(row=applicant + 2, column=reviewer + 2, value=str(rank+1))

        # do weight sheet stuff
        weight_sheet = workbook.create_sheet("weights")
        self.add_headers_to_sheet(weight_sheet, output_matrix)
        
        # fill weight matrix
        for j, applicant in enumerate(output_matrix.applicants):
            for i, reviewer in enumerate(output_matrix.reviewers):
                weight_sheet.cell(row=j + 2, column=i + 2, value=str(input_matrix.matrix[i][j]))

        # highlight matches in matrix
        goldfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00FFD700'))
        silverfill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00D8D8D8'))
        bronzefill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00CD7F32'))
        for applicant, assignments in enumerate(output_matrix.applicant_assignments):
            #print(assignments)
            for rank, reviewer in enumerate(assignments):
                cell = weight_sheet.cell(row=applicant + 2, column=reviewer + 2)
                if (rank == 0):
                    cell.fill = goldfill
                elif rank == 1:
                    cell.fill = silverfill
                else:
                    cell.fill = bronzefill

        # save xlsx file
        # get location to save file
        save_location = fd.asksaveasfilename(filetypes=[("Microsoft Excel spreadsheet", "*.xlsx")], defaultextension=[("Microsoft Excel spreadsheet", "*.xlsx")])
        if (save_location == None):
            return  # return if no location is set
        # attempt to save
        workbook.save(save_location)
        return
    
    def add_headers_to_sheet(self, sheet, output_matrix):
        # setup reviewer and applicant names
        for i, reviewer in enumerate(output_matrix.reviewers):
            sheet.cell(row=1, column=i+2, value=reviewer)
        for i, applicant in enumerate(output_matrix.applicants):
            sheet.cell(row=i+2, column=1, value=applicant)
    
    

    def matrix_to_sheet_coords(self, reviewer: int, applicant: int) -> tuple[int, int]:
        # topleft position is (2,2)
        return (applicant + 2, reviewer+2)
    
    def add_UI(self, ui: UIInterface):
        self.ui = ui
        return